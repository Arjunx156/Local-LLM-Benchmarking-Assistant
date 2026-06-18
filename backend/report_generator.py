"""
report_generator.py — Excel (openpyxl + kaleido) and JSON report generation.

Excel structure:
  Sheet 1: Summary  (model rankings + system specs)
  Sheet 2–N: Per-model question-by-question breakdown
  Sheet N+1: Charts  (embedded PNG images)

JSON structure: full raw results + metadata.
"""
from __future__ import annotations

import json
import platform
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List

import pandas as pd
import plotly.graph_objects as go
import psutil

from backend.config import settings
from backend.job_store import BenchmarkJob, QuestionResult

# Professional color palette (consistent across all charts)
PALETTE = ["#6366f1", "#f59e0b", "#10b981", "#ef4444", "#3b82f6", "#8b5cf6", "#ec4899"]
CHART_BG = "#0f172a"
CHART_PAPER_BG = "#1e293b"
CHART_FONT_COLOR = "#f8fafc"
CHART_GRID_COLOR = "#334155"


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _results_to_df(results: List[QuestionResult]) -> pd.DataFrame:
    if not results:
        return pd.DataFrame()
    rows = [
        {
            "model": r.model,
            "question_id": r.question_id,
            "category": r.category,
            "difficulty": r.difficulty,
            "score": r.score,
            "ttft_ms": r.ttft_ms,
            "total_time_ms": r.total_time_ms,
            "tokens_per_second": r.tokens_per_second,
            "tokens_generated": r.tokens_generated,
            "peak_ram_mb": r.peak_ram_mb,
            "avg_cpu_percent": r.avg_cpu_percent,
            "evaluation_method": r.evaluation_method,
            "error": r.error or "",
        }
        for r in results
    ]
    return pd.DataFrame(rows)


def _model_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    grp = df.groupby("model").agg(
        avg_score=("score", "mean"),
        avg_tok_per_sec=("tokens_per_second", "mean"),
        avg_latency_ms=("total_time_ms", "mean"),
        avg_ttft_ms=("ttft_ms", "mean"),
        avg_ram_mb=("peak_ram_mb", "mean"),
        avg_cpu_pct=("avg_cpu_percent", "mean"),
        total_questions=("question_id", "count"),
    ).reset_index()
    grp["avg_score_pct"] = (grp["avg_score"] * 100).round(1)
    return grp.sort_values("avg_score", ascending=False)


def _system_metadata() -> dict:
    vm = psutil.virtual_memory()
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "os": f"{platform.system()} {platform.release()}",
        "cpu": platform.processor() or "Unknown",
        "ram_total_gb": round(vm.total / 1024**3, 2),
        "python_version": platform.python_version(),
    }


# ─── Chart builders ───────────────────────────────────────────────────────────

def _chart_layout(title: str) -> dict:
    return dict(
        title=dict(text=title, font=dict(color=CHART_FONT_COLOR, size=14)),
        paper_bgcolor=CHART_PAPER_BG,
        plot_bgcolor=CHART_BG,
        font=dict(color=CHART_FONT_COLOR),
        xaxis=dict(gridcolor=CHART_GRID_COLOR, zerolinecolor=CHART_GRID_COLOR),
        yaxis=dict(gridcolor=CHART_GRID_COLOR, zerolinecolor=CHART_GRID_COLOR),
        margin=dict(l=60, r=30, t=50, b=60),
    )


def build_bar_chart(df: pd.DataFrame) -> go.Figure:
    """Tokens/second per model — horizontal bar chart."""
    summary = _model_summary(df)
    fig = go.Figure(go.Bar(
        x=summary["avg_tok_per_sec"].round(1),
        y=summary["model"],
        orientation="h",
        marker_color=PALETTE[:len(summary)],
        text=summary["avg_tok_per_sec"].round(1),
        textposition="outside",
    ))
    fig.update_layout(**_chart_layout("⚡ Avg Tokens / Second by Model"))
    return fig


def build_radar_chart(df: pd.DataFrame) -> go.Figure:
    """Radar chart — model scores across benchmark categories."""
    categories = df["category"].unique().tolist()
    fig = go.Figure()
    for i, model in enumerate(df["model"].unique()):
        model_df = df[df["model"] == model]
        cat_scores = [
            model_df[model_df["category"] == c]["score"].mean() * 100
            for c in categories
        ]
        cat_scores_closed = cat_scores + [cat_scores[0]]
        categories_closed = categories + [categories[0]]
        fig.add_trace(go.Scatterpolar(
            r=cat_scores_closed,
            theta=categories_closed,
            fill="toself",
            name=model,
            line_color=PALETTE[i % len(PALETTE)],
            fillcolor=PALETTE[i % len(PALETTE)],
            opacity=0.3,
        ))
    fig.update_layout(
        polar=dict(
            bgcolor=CHART_BG,
            radialaxis=dict(
                visible=True, range=[0, 100],
                gridcolor=CHART_GRID_COLOR, color=CHART_FONT_COLOR,
            ),
            angularaxis=dict(color=CHART_FONT_COLOR),
        ),
        **{k: v for k, v in _chart_layout("🎯 Score Radar by Category").items()
           if k not in ("xaxis", "yaxis", "plot_bgcolor")},
    )
    return fig


def build_latency_chart(df: pd.DataFrame) -> go.Figure:
    """Line chart — avg latency vs difficulty per model."""
    difficulty_order = {"easy": 1, "medium": 2, "hard": 3}
    df = df.copy()
    df["diff_order"] = df["difficulty"].map(difficulty_order)
    fig = go.Figure()
    for i, model in enumerate(df["model"].unique()):
        mdf = (
            df[df["model"] == model]
            .groupby(["difficulty", "diff_order"])["total_time_ms"]
            .mean()
            .reset_index()
            .sort_values("diff_order")
        )
        fig.add_trace(go.Scatter(
            x=mdf["difficulty"],
            y=mdf["total_time_ms"].round(0),
            name=model,
            mode="lines+markers",
            line=dict(color=PALETTE[i % len(PALETTE)], width=2),
            marker=dict(size=8),
        ))
    fig.update_layout(**_chart_layout("⏱️ Avg Latency (ms) vs Difficulty"))
    fig.update_yaxes(title_text="Latency (ms)")
    return fig


def build_heatmap_chart(df: pd.DataFrame) -> go.Figure:
    """Heatmap — model × category score matrix."""
    pivot = df.pivot_table(
        values="score", index="model", columns="category", aggfunc="mean"
    ).round(2)
    fig = go.Figure(go.Heatmap(
        z=pivot.values * 100,
        x=pivot.columns.tolist(),
        y=pivot.index.tolist(),
        colorscale=[[0, "#ef4444"], [0.5, "#f59e0b"], [1, "#10b981"]],
        zmin=0, zmax=100,
        text=(pivot.values * 100).round(1),
        texttemplate="%{text}%",
        colorbar=dict(title="Score %", tickfont=dict(color=CHART_FONT_COLOR)),
    ))
    fig.update_layout(**{k: v for k, v in _chart_layout("🔥 Score Heatmap (Model × Category)").items()
                        if k not in ("xaxis", "yaxis")})
    return fig


# ─── PNG rendering ────────────────────────────────────────────────────────────

def _fig_to_png_bytes(fig: go.Figure, width: int = 900, height: int = 500) -> bytes:
    """Render plotly figure to PNG. Tries kaleido first, falls back gracefully."""
    try:
        return fig.to_image(format="png", width=width, height=height, scale=2)
    except Exception:
        # kaleido not installed or failed — return empty bytes; chart will be skipped
        return b""


# ─── Excel report ─────────────────────────────────────────────────────────────

def generate_excel_report(job: BenchmarkJob) -> Path:
    """
    Build a multi-sheet Excel report and return the file path.
    """
    df = _results_to_df(job.results)
    summary_df = _model_summary(df)
    meta = _system_metadata()

    out_dir = settings.REPORTS_DIR / job.job_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "report.xlsx"

    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        # ── Sheet 1: Summary ──────────────────────────────────────────────────
        summary_display = summary_df.rename(columns={
            "model": "Model",
            "avg_score_pct": "Avg Score (%)",
            "avg_tok_per_sec": "Avg Tok/s",
            "avg_latency_ms": "Avg Latency (ms)",
            "avg_ttft_ms": "Avg TTFT (ms)",
            "avg_ram_mb": "Avg Peak RAM (MB)",
            "avg_cpu_pct": "Avg CPU (%)",
            "total_questions": "Questions",
        }).round(2)
        summary_display.to_excel(writer, sheet_name="Summary", index=False, startrow=6)

        ws = writer.sheets["Summary"]
        ws["A1"] = "LLM Benchmarking Report"
        ws["A2"] = f"Suite: {job.suite}  |  Models: {', '.join(job.models)}"
        ws["A3"] = f"Generated: {meta['generated_at']}"
        ws["A4"] = f"System: {meta['cpu']} | {meta['ram_total_gb']} GB RAM | {meta['os']}"

        # ── Per-model sheets ───────────────────────────────────────────────────
        for model in job.models:
            model_df = df[df["model"] == model].copy()
            if model_df.empty:
                continue
            sheet_name = model.replace(":", "_")[:31]   # Excel limit
            model_display = model_df[[
                "question_id", "category", "difficulty",
                "score", "total_time_ms", "ttft_ms",
                "tokens_per_second", "tokens_generated",
                "peak_ram_mb", "avg_cpu_percent", "error",
            ]].rename(columns={
                "question_id": "Q ID", "category": "Category",
                "difficulty": "Difficulty", "score": "Score (0-1)",
                "total_time_ms": "Total Time (ms)", "ttft_ms": "TTFT (ms)",
                "tokens_per_second": "Tok/s", "tokens_generated": "Tokens",
                "peak_ram_mb": "Peak RAM (MB)", "avg_cpu_percent": "CPU (%)",
            })
            model_display.to_excel(writer, sheet_name=sheet_name, index=False)

        # ── Charts sheet ──────────────────────────────────────────────────────
        if not df.empty:
            try:
                from openpyxl.drawing.image import Image as XLImage
                import io

                writer.book.create_sheet("Charts")
                ws_charts = writer.book["Charts"]
                ws_charts["A1"] = "Benchmark Charts"

                charts = [
                    ("bar", build_bar_chart(df)),
                    ("radar", build_radar_chart(df)),
                    ("latency", build_latency_chart(df)),
                    ("heatmap", build_heatmap_chart(df)),
                ]
                row_offsets = [3, 30, 57, 84]
                for (name, fig), row in zip(charts, row_offsets):
                    png = _fig_to_png_bytes(fig)
                    if not png:
                        continue
                    img = XLImage(io.BytesIO(png))
                    img.width = 700
                    img.height = 380
                    ws_charts.add_image(img, f"A{row}")
            except Exception:
                pass   # Charts are optional

    return out_path


# ─── JSON report ──────────────────────────────────────────────────────────────

def generate_json_report(job: BenchmarkJob) -> Path:
    """Build a comprehensive JSON report and return the file path."""
    df = _results_to_df(job.results)
    summary = _model_summary(df).to_dict(orient="records") if not df.empty else []
    meta = _system_metadata()

    report = {
        "metadata": {
            **meta,
            "job_id": job.job_id,
            "suite": job.suite,
            "models": job.models,
            "suite_version": "1.0",
        },
        "summary": summary,
        "results_by_model": {},
    }

    for model in job.models:
        model_results = [r for r in job.results if r.model == model]
        report["results_by_model"][model] = [
            {
                "question_id": r.question_id,
                "category": r.category,
                "difficulty": r.difficulty,
                "prompt": r.prompt,
                "expected_answer": r.expected_answer,
                "model_response": r.response_text,
                "evaluation_method": r.evaluation_method,
                "score": r.score,
                "ttft_ms": r.ttft_ms,
                "total_time_ms": r.total_time_ms,
                "tokens_per_second": r.tokens_per_second,
                "tokens_generated": r.tokens_generated,
                "peak_ram_mb": r.peak_ram_mb,
                "avg_cpu_percent": r.avg_cpu_percent,
                "error": r.error,
            }
            for r in model_results
        ]

    out_dir = settings.REPORTS_DIR / job.job_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "report.json"
    out_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    return out_path
